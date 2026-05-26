from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter


def clear_sheet_except_header(ws):
    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)


def write_headers_if_missing(ws, headers: list[str]):
    for idx, header in enumerate(headers, start=1):
        ws.cell(row=1, column=idx, value=header)
        ws.cell(row=1, column=idx).font = Font(bold=True)
        ws.cell(row=1, column=idx).alignment = Alignment(horizontal="center", vertical="center")


def write_news_sheet(workbook_path: str, rows: list[dict]):
    wb = load_workbook(workbook_path)
    ws = wb["뉴스 수집"]

    headers = ["키워드", "검색어", "순위", "기사제목", "날짜", "요약", "기사링크"]
    write_headers_if_missing(ws, headers)
    clear_sheet_except_header(ws)

    for row_idx, row in enumerate(rows, start=2):
        ws.cell(row=row_idx, column=1, value=row["키워드"])
        ws.cell(row=row_idx, column=2, value=row["검색어"])
        ws.cell(row=row_idx, column=3, value=row["순위"])
        ws.cell(row=row_idx, column=4, value=row["기사제목"])
        ws.cell(row=row_idx, column=5, value=row["날짜"])
        ws.cell(row=row_idx, column=6, value=row["요약"])
        ws.cell(row=row_idx, column=7, value=row["기사링크"])

    format_sheet(ws)
    wb.save(workbook_path)


def write_trend_sheet(workbook_path: str, rows: list[dict]):
    wb = load_workbook(workbook_path)
    ws = wb["트렌드 요약"]

    headers = ["키워드", "검색어", "기사건수", "종합 트렌드/최신 동향"]
    write_headers_if_missing(ws, headers)
    clear_sheet_except_header(ws)

    for row_idx, row in enumerate(rows, start=2):
        ws.cell(row=row_idx, column=1, value=row["키워드"])
        ws.cell(row=row_idx, column=2, value=row["검색어"])
        ws.cell(row=row_idx, column=3, value=row["기사건수"])
        ws.cell(row=row_idx, column=4, value=row["종합 트렌드/최신 동향"])

    format_sheet(ws)
    wb.save(workbook_path)


def format_sheet(ws):
    ws.freeze_panes = "A2"

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    for column_cells in ws.columns:
        max_length = 0
        col_letter = get_column_letter(column_cells[0].column)

        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            if len(value) > max_length:
                max_length = len(value)

        if col_letter in ["D", "F"]:
            ws.column_dimensions[col_letter].width = 45
        elif col_letter == "G":
            ws.column_dimensions[col_letter].width = 60
        else:
            ws.column_dimensions[col_letter].width = min(max_length + 3, 25)