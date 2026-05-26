from pathlib import Path
from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Font, Alignment


KEYWORDS = [
    "반도체",
    "철강",
    "채용",
    "세라믹",
    "고분자",
    "디스플레이",
    "고려대학교",
]


def style_header(ws):
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")


def create_config_excel():
    output_dir = Path("config")
    output_dir.mkdir(parents=True, exist_ok=True)
    output_path = output_dir / "keyword_config.xlsx"

    wb = Workbook()

    ws_setting = wb.active
    ws_setting.title = "설정"
    ws_setting["A1"] = "키워드"
    ws_setting["B1"] = "사용여부"
    style_header(ws_setting)

    for idx, keyword in enumerate(KEYWORDS, start=2):
        ws_setting[f"A{idx}"] = keyword
        ws_setting[f"B{idx}"] = "N"

    dv = DataValidation(type="list", formula1='"Y,N"', allow_blank=False)
    ws_setting.add_data_validation(dv)
    dv.add(f"B2:B{len(KEYWORDS)+1}")

    ws_setting.column_dimensions["A"].width = 18
    ws_setting.column_dimensions["B"].width = 12
    ws_setting.freeze_panes = "A2"

    ws_news = wb.create_sheet("뉴스 수집")
    news_headers = ["키워드", "검색어", "순위", "기사제목", "날짜", "요약", "기사링크"]
    for col_idx, header in enumerate(news_headers, start=1):
        ws_news.cell(row=1, column=col_idx, value=header)
    style_header(ws_news)
    ws_news.freeze_panes = "A2"

    ws_trend = wb.create_sheet("트렌드 요약")
    trend_headers = ["키워드", "검색어", "기사건수", "종합 트렌드/최신 동향"]
    for col_idx, header in enumerate(trend_headers, start=1):
        ws_trend.cell(row=1, column=col_idx, value=header)
    style_header(ws_trend)
    ws_trend.freeze_panes = "A2"

    wb.save(output_path)
    print(f"생성 완료: {output_path.resolve()}")


if __name__ == "__main__":
    create_config_excel()