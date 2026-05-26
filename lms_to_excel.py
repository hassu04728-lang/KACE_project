from urllib.request import urlopen
from datetime import datetime, timezone, timedelta
import csv
import re
import time

ICS_URL = "https://mylms.korea.ac.kr/feeds/calendars/user_pEVsrl3po0EqMbTTlvvjUbup6MczkhL6VbvR1ZXk.ics"
ANNOUNCEMENTS_URL = "https://mylms.korea.ac.kr/accounts/1/external_tools/10?launch_type=global_navigation"

OUTPUT_CSV = "lms_calendar.csv"
OUTPUT_XLSX = "lms_calendar.xlsx"


# =========================
# 1) 캘린더(.ics) 처리
# =========================
def download_ics(url: str) -> str:
    with urlopen(url) as response:
        return response.read().decode("utf-8", errors="replace")


def unfold_ics_lines(text: str):
    raw_lines = text.replace("\r\n", "\n").replace("\r", "\n").split("\n")
    lines = []

    for line in raw_lines:
        if (line.startswith(" ") or line.startswith("\t")) and lines:
            lines[-1] += line[1:]
        else:
            lines.append(line)

    return lines


def unescape_ics_text(value: str) -> str:
    return (
        value.replace("\\n", "\n")
        .replace("\\N", "\n")
        .replace("\\,", ",")
        .replace("\\;", ";")
        .replace("\\\\", "\\")
    )


def parse_datetime(key_part: str, value: str):
    value = value.strip()

    if "VALUE=DATE" in key_part or (len(value) == 8 and value.isdigit()):
        try:
            return datetime.strptime(value, "%Y%m%d")
        except ValueError:
            return None

    if value.endswith("Z"):
        for fmt in ("%Y%m%dT%H%M%SZ", "%Y%m%dT%H%MZ"):
            try:
                dt = datetime.strptime(value, fmt)
                return dt.replace(tzinfo=timezone.utc)
            except ValueError:
                continue
        return None

    for fmt in ("%Y%m%dT%H%M%S", "%Y%m%dT%H%M"):
        try:
            return datetime.strptime(value, fmt)
        except ValueError:
            continue

    return None


def format_datetime(dt):
    if dt is None:
        return ""

    if dt.tzinfo is not None:
        kst = timezone(timedelta(hours=9))
        dt = dt.astimezone(kst)

    return dt.strftime("%Y-%m-%d %H:%M")


def detect_type(title: str, description: str) -> str:
    text = f"{title} {description}".lower()

    due_keywords = [
        "due", "assignment", "quiz", "exam", "test", "submission",
        "과제", "마감", "퀴즈", "시험", "제출"
    ]
    material_keywords = [
        "material", "file", "lecture note", "slides", "handout",
        "자료", "강의자료", "수업자료", "pdf", "lecture"
    ]

    if any(keyword in text for keyword in due_keywords):
        return "과제/마감"
    if any(keyword in text for keyword in material_keywords):
        return "강의자료/수업"
    return "기타 일정"


def parse_ics_events(ics_text: str):
    lines = unfold_ics_lines(ics_text)
    events = []
    current_event = None

    for line in lines:
        if line == "BEGIN:VEVENT":
            current_event = {}
            continue

        if line == "END:VEVENT":
            if current_event:
                title = unescape_ics_text(current_event.get("SUMMARY", ""))
                description = unescape_ics_text(current_event.get("DESCRIPTION", ""))
                url = current_event.get("URL", "")

                start_dt = current_event.get("_START_DT")
                end_dt = current_event.get("_END_DT")

                event_type = detect_type(title, description)

                if event_type == "과제/마감":
                    date_str = format_datetime(start_dt or end_dt)
                else:
                    date_str = format_datetime(end_dt or start_dt)

                events.append({
                    "유형": event_type,
                    "제목": title,
                    "종료/마감": date_str,
                    "설명": description,
                    "URL": url
                })

            current_event = None
            continue

        if current_event is None:
            continue

        if ":" not in line:
            continue

        key_part, value = line.split(":", 1)
        key_name = key_part.split(";", 1)[0].upper()

        if key_name == "DTSTART":
            current_event["_START_DT"] = parse_datetime(key_part, value)
        elif key_name == "DTEND":
            current_event["_END_DT"] = parse_datetime(key_part, value)
        else:
            current_event[key_name] = value

    def sort_key(event):
        date_str = event["종료/마감"]
        if not date_str:
            return datetime.max

        for fmt in ("%Y-%m-%d %H:%M", "%Y-%m-%d"):
            try:
                return datetime.strptime(date_str, fmt)
            except ValueError:
                continue

        return datetime.max

    events.sort(key=sort_key)
    return events


def save_to_csv(events, filename):
    if not events:
        print("저장할 일정이 없습니다.")
        return

    fieldnames = ["유형", "제목", "종료/마감", "설명", "URL"]

    with open(filename, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(events)

    print(f"CSV 저장 완료: {filename}")


# =========================
# 2) 공지사항 처리
# =========================
def extract_course_and_title(raw_title: str):
    raw_title = " ".join((raw_title or "").split())
    match = re.match(r"^\[([^\]]+)\](.*)$", raw_title)

    if match:
        return match.group(1).strip(), match.group(2).strip()

    return "", raw_title.strip()


def build_driver():
    from selenium import webdriver
    from selenium.webdriver.chrome.options import Options

    options = Options()
    options.add_argument("--start-maximized")
    options.add_experimental_option("detach", True)

    return webdriver.Chrome(options=options)


def switch_to_frame_with_table(driver, timeout=60):
    from selenium.webdriver.common.by import By

    end_time = time.time() + timeout

    while time.time() < end_time:
        try:
            driver.switch_to.default_content()
            rows = driver.find_elements(By.CSS_SELECTOR, "table tbody tr")
            if rows:
                return True
        except Exception:
            pass

        try:
            driver.switch_to.default_content()
            frames = driver.find_elements(By.CSS_SELECTOR, "iframe, frame")

            for i in range(len(frames)):
                try:
                    driver.switch_to.default_content()
                    frames = driver.find_elements(By.CSS_SELECTOR, "iframe, frame")
                    driver.switch_to.frame(frames[i])

                    rows = driver.find_elements(By.CSS_SELECTOR, "table tbody tr")
                    if rows:
                        return True
                except Exception:
                    continue
        except Exception:
            pass

        time.sleep(1)

    driver.switch_to.default_content()
    return False


def extract_announcements_from_table(driver):
    from selenium.webdriver.common.by import By

    announcements = []
    seen = set()

    rows = driver.find_elements(By.CSS_SELECTOR, "table tbody tr")

    for row in rows:
        tds = row.find_elements(By.TAG_NAME, "td")

        # 표 구조:
        # 0: No / 1: 제목 / 2: 작성자 / 3: 등록일시 / 4: 조회수
        if len(tds) < 5:
            continue

        try:
            title_el = tds[1].find_element(By.TAG_NAME, "a")
            raw_title = " ".join(title_el.text.split())
        except Exception:
            raw_title = " ".join(tds[1].text.split())

        date_text = " ".join(tds[3].text.split())

        if not raw_title:
            continue

        course, clean_title = extract_course_and_title(raw_title)

        item = {
            "과목": course,
            "제목": clean_title,
            "등록 일시": date_text,
        }

        key = (item["과목"], item["제목"], item["등록 일시"])
        if key in seen:
            continue

        seen.add(key)
        announcements.append(item)

    return announcements


def scrape_announcements():
    driver = build_driver()

    try:
        print("공지사항 페이지 여는 중...")
        driver.get(ANNOUNCEMENTS_URL)

        input(
            "\n브라우저가 열렸습니다.\n"
            "로그인 후, '전체 게시물 보기' 표가 실제로 화면에 보이게 한 다음\n"
            "터미널에서 Enter를 누르세요..."
        )

        print("표가 있는 문서/iframe 찾는 중...")
        ok = switch_to_frame_with_table(driver, timeout=60)

        if not ok:
            driver.save_screenshot("announcement_debug.png")
            with open("announcement_debug.html", "w", encoding="utf-8") as f:
                f.write(driver.page_source)

            print("공지사항 표를 찾지 못했습니다.")
            print("디버그 파일 저장: announcement_debug.png, announcement_debug.html")
            return []

        announcements = extract_announcements_from_table(driver)
        print(f"공지사항 {len(announcements)}개를 찾았습니다.")
        return announcements

    except Exception as e:
        try:
            driver.save_screenshot("announcement_error.png")
            with open("announcement_error.html", "w", encoding="utf-8") as f:
                f.write(driver.page_source)
            print("에러 시점 파일 저장: announcement_error.png, announcement_error.html")
        except Exception:
            pass

        print("공지사항 수집 중 오류:", repr(e))
        input("확인 후 Enter를 누르세요...")
        return []

    finally:
        try:
            driver.quit()
        except Exception:
            pass


# =========================
# 3) 엑셀 저장
# =========================
def style_sheet(ws, headers, widths):
    from openpyxl.styles import Font, PatternFill, Alignment

    ws.append(headers)

    header_fill = PatternFill("solid", fgColor="D9EAF7")
    header_font = Font(bold=True)

    for col_idx, _ in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def save_to_xlsx(events, announcements, filename):
    try:
        from openpyxl import Workbook
    except ImportError:
        print("openpyxl이 설치되지 않아 XLSX 저장은 건너뜁니다.")
        print("설치 명령어: py -m pip install openpyxl")
        return

    wb = Workbook()

    ws1 = wb.active
    ws1.title = "LMS 일정"

    event_headers = ["유형", "제목", "종료/마감", "설명", "URL"]
    style_sheet(
        ws1,
        event_headers,
        {"A": 14, "B": 40, "C": 18, "D": 70, "E": 55},
    )

    for event in events:
        ws1.append([
            event["유형"],
            event["제목"],
            event["종료/마감"],
            event["설명"],
            event["URL"],
        ])

    ws2 = wb.create_sheet("공지사항")
    ann_headers = ["과목", "제목", "등록 일시"]
    style_sheet(
        ws2,
        ann_headers,
        {"A": 45, "B": 85, "C": 20},
    )

    for item in announcements:
        ws2.append([
            item["과목"],
            item["제목"],
            item["등록 일시"],
        ])

    wb.save(filename)
    print(f"XLSX 저장 완료: {filename}")


# =========================
# 4) 실행
# =========================
def main():
    print("LMS 캘린더 다운로드 중...")
    ics_text = download_ics(ICS_URL)

    print("일정 파싱 중...")
    events = parse_ics_events(ics_text)
    print(f"총 {len(events)}개의 일정 발견")

    save_to_csv(events, OUTPUT_CSV)

    print("공지사항 수집 시작...")
    announcements = scrape_announcements()
    print(f"총 {len(announcements)}개의 공지사항 발견")

    save_to_xlsx(events, announcements, OUTPUT_XLSX)

    print("완료!")
    input("엔터를 누르면 프로그램이 끝납니다...")


if __name__ == "__main__":
    main()