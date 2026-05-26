from urllib.request import Request, urlopen
from urllib.parse import urljoin
from datetime import datetime, timezone, timedelta
import html
import sys
import json
import time
import re
import csv
from pathlib import Path

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.options import Options
from selenium.common.exceptions import WebDriverException


# =========================
# 0) 기본 설정
# =========================
BASE_DIR = Path(__file__).resolve().parent

DEBUG_FILE = BASE_DIR / "login_debug.txt"
OUTPUT_XLSX = BASE_DIR / "lms_calendar.xlsx"
OUTPUT_CSV = BASE_DIR / "lms_calendar.csv"

ANNOUNCEMENTS_URL = "https://mylms.korea.ac.kr/accounts/1/external_tools/10?launch_type=global_navigation"
CALENDAR_URL = "https://mylms.korea.ac.kr/calendar#view_name=month&view_start=2026-05-19"

KEEP_BROWSER_OPEN = True


def log(message):
    with open(DEBUG_FILE, "a", encoding="utf-8") as f:
        f.write(str(message) + "\n")


def read_payload():
    raw = sys.stdin.read().strip()
    if not raw:
        raise RuntimeError("VBA에서 전달된 로그인 정보가 없습니다.")
    return json.loads(raw)


# =========================
# 1) 로그인 처리
# =========================
def is_visible_input(element):
    try:
        if not element.is_displayed():
            return False
        if not element.is_enabled():
            return False
        if element.get_attribute("disabled"):
            return False
        if element.get_attribute("readonly"):
            return False
        return True
    except WebDriverException:
        return False


def safe_type(element, text):
    element.click()
    time.sleep(0.2)
    element.send_keys(Keys.CONTROL, "a")
    element.send_keys(Keys.DELETE)
    time.sleep(0.1)
    element.send_keys(text)


def dump_inputs(driver, title):
    log("")
    log("=" * 60)
    log(title)
    log(f"URL: {driver.current_url}")
    log(f"TITLE: {driver.title}")

    inputs = driver.find_elements(By.CSS_SELECTOR, "input")
    log(f"input 개수: {len(inputs)}")

    for i, inp in enumerate(inputs, start=1):
        try:
            log(
                f"{i}. "
                f"type={inp.get_attribute('type')}, "
                f"id={inp.get_attribute('id')}, "
                f"name={inp.get_attribute('name')}, "
                f"placeholder={inp.get_attribute('placeholder')}, "
                f"class={inp.get_attribute('class')}, "
                f"displayed={inp.is_displayed()}, "
                f"enabled={inp.is_enabled()}"
            )
        except Exception as e:
            log(f"{i}. input 정보 읽기 실패: {e}")


def click_login_or_enter(driver, password_box):
    """
    로그인 버튼을 최대한 확실하게 누른다.
    1) 비밀번호 input이 들어있는 form 안의 submit 버튼 클릭
    2) 화면 전체에서 로그인 버튼 추정 요소 클릭
    3) 비밀번호 칸에서 Enter
    4) 마지막으로 form.requestSubmit()
    """

    # 1. password input이 포함된 form 내부 버튼 먼저 찾기
    form = None
    try:
        form = password_box.find_element(By.XPATH, "./ancestor::form[1]")

        button_selectors = [
            "button[type='submit']",
            "input[type='submit']",
            "input[type='button']",
            "button",
            "a",
            "input[type='image']",
        ]

        for selector in button_selectors:
            buttons = form.find_elements(By.CSS_SELECTOR, selector)

            for btn in buttons:
                try:
                    text = (btn.text or "").strip()
                    value = (btn.get_attribute("value") or "").strip()
                    title = (btn.get_attribute("title") or "").strip()
                    aria = (btn.get_attribute("aria-label") or "").strip()
                    onclick = (btn.get_attribute("onclick") or "").strip()
                    btn_id = (btn.get_attribute("id") or "").strip()
                    btn_name = (btn.get_attribute("name") or "").strip()
                    btn_class = (btn.get_attribute("class") or "").strip()

                    combined = f"{text} {value} {title} {aria} {onclick} {btn_id} {btn_name} {btn_class}".lower()

                    if (
                        "로그인" in combined
                        or "login" in combined
                        or "submit" in combined
                        or selector in ["button[type='submit']", "input[type='submit']", "input[type='image']"]
                    ):
                        driver.execute_script(
                            "arguments[0].scrollIntoView({block:'center'});",
                            btn
                        )
                        time.sleep(0.2)
                        driver.execute_script("arguments[0].click();", btn)
                        log("form 내부 로그인 버튼 클릭 성공")
                        return True

                except Exception:
                    continue

    except Exception as e:
        log(f"form 내부 로그인 버튼 탐색 실패: {e}")

    # 2. 현재 frame 전체에서 로그인 버튼 추정 요소 JS로 클릭
    try:
        clicked = driver.execute_script("""
            const candidates = Array.from(document.querySelectorAll(
                "button, input[type='submit'], input[type='button'], input[type='image'], a, div[role='button'], span[role='button']"
            ));

            for (const el of candidates) {
                const text = [
                    el.innerText,
                    el.value,
                    el.title,
                    el.getAttribute("aria-label"),
                    el.getAttribute("onclick"),
                    el.id,
                    el.name,
                    el.className,
                    el.src
                ].filter(Boolean).join(" ").toLowerCase();

                if (
                    text.includes("로그인") ||
                    text.includes("login") ||
                    text.includes("submit")
                ) {
                    el.scrollIntoView({block: "center"});
                    el.click();
                    return text;
                }
            }

            return null;
        """)

        if clicked:
            log(f"JS로 로그인 버튼 클릭 성공: {clicked}")
            return True

    except Exception as e:
        log(f"JS 로그인 버튼 클릭 실패: {e}")

    # 3. 비밀번호 칸에서 Enter
    try:
        password_box.click()
        time.sleep(0.2)
        password_box.send_keys(Keys.ENTER)
        log("비밀번호 칸에서 Enter 입력")
        time.sleep(1)
        return True

    except Exception as e:
        log(f"Enter 입력 실패: {e}")

    # 4. 마지막 수단: form.requestSubmit()
    try:
        if form is not None:
            driver.execute_script("""
                const form = arguments[0];
                if (form.requestSubmit) {
                    form.requestSubmit();
                } else {
                    form.submit();
                }
            """, form)
            log("form.requestSubmit 실행")
            return True

    except Exception as e:
        log(f"form.requestSubmit 실패: {e}")

    return False


def try_fill_current_context(driver, user_id, password):
    dump_inputs(driver, "현재 frame에서 input 탐색")

    inputs = driver.find_elements(By.CSS_SELECTOR, "input")
    visible_inputs = [x for x in inputs if is_visible_input(x)]

    password_boxes = []
    user_boxes = []

    for inp in visible_inputs:
        input_type = (inp.get_attribute("type") or "").lower()
        input_name = (inp.get_attribute("name") or "").lower()
        input_id = (inp.get_attribute("id") or "").lower()
        placeholder = (inp.get_attribute("placeholder") or "").lower()

        if input_type == "password":
            password_boxes.append(inp)
            continue

        if input_type in ["text", "email", "tel", "search", ""]:
            user_boxes.append(inp)
            continue

        if any(key in input_name + input_id + placeholder for key in ["id", "user", "login"]):
            user_boxes.append(inp)

    if not password_boxes:
        log("비밀번호 input을 찾지 못함")
        return False

    if not user_boxes:
        log("아이디 input을 찾지 못함")
        return False

    user_box = user_boxes[0]
    password_box = password_boxes[0]

    safe_type(user_box, user_id)
    log("아이디 입력 성공")

    safe_type(password_box, password)
    log("비밀번호 입력 성공")

    click_login_or_enter(driver, password_box)
    return True


def recursive_find_and_fill(driver, user_id, password, depth=0, max_depth=3):
    if depth > max_depth:
        return False

    if try_fill_current_context(driver, user_id, password):
        return True

    frames = driver.find_elements(By.CSS_SELECTOR, "iframe, frame")
    log(f"depth={depth}, frame 개수: {len(frames)}")

    for i in range(len(frames)):
        try:
            frames = driver.find_elements(By.CSS_SELECTOR, "iframe, frame")
            driver.switch_to.frame(frames[i])
            log(f"frame {i} 진입")

            if recursive_find_and_fill(driver, user_id, password, depth + 1, max_depth):
                return True

            driver.switch_to.parent_frame()
            log(f"frame {i} 빠져나옴")

        except Exception as e:
            log(f"frame {i} 처리 실패: {e}")
            try:
                driver.switch_to.default_content()
            except Exception:
                pass

    return False


def wait_after_login(driver, timeout=50):
    """
    로그인 후 URL이 바뀌는 것은 정상.
    여기서는 로그인된 세션이 생길 때까지 조금 기다린다.
    """
    end_time = time.time() + timeout

    while time.time() < end_time:
        current_url = driver.current_url.lower()

        if "mylms.korea.ac.kr" in current_url or "lms.korea.ac.kr" in current_url:
            log(f"로그인 후 LMS 이동 확인: {driver.current_url}")
            time.sleep(3)
            return True

        time.sleep(1)

    log(f"로그인 후 LMS URL 확인 실패. 현재 URL: {driver.current_url}")
    return False


# =========================
# 2) 공지사항 처리
# =========================
def extract_course_and_title(raw_title: str):
    raw_title = " ".join((raw_title or "").split())
    match = re.match(r"^\[([^\]]+)\](.*)$", raw_title)

    if match:
        return match.group(1).strip(), match.group(2).strip()

    return "", raw_title.strip()


def click_all_posts_if_exists(driver):
    """
    공지사항 페이지에서 '전체 게시물 보기' 같은 버튼이 있으면 누른다.
    없으면 그냥 넘어간다.
    """
    candidates = driver.find_elements(By.CSS_SELECTOR, "a, button, input[type='button'], input[type='submit']")

    for el in candidates:
        try:
            text = " ".join((el.text or "").split())
            value = " ".join((el.get_attribute("value") or "").split())
            title = " ".join((el.get_attribute("title") or "").split())
            combined = f"{text} {value} {title}"

            if "전체" in combined and ("게시물" in combined or "공지" in combined):
                if el.is_displayed() and el.is_enabled():
                    el.click()
                    log("'전체 게시물 보기' 추정 버튼 클릭")
                    time.sleep(3)
                    return True
        except Exception:
            continue

    return False


def switch_to_frame_with_table(driver, timeout=60):
    end_time = time.time() + timeout

    while time.time() < end_time:
        try:
            driver.switch_to.default_content()
            click_all_posts_if_exists(driver)

            rows = driver.find_elements(By.CSS_SELECTOR, "table tbody tr")
            if rows:
                return True
        except Exception:
            pass

        try:
            driver.switch_to.default_content()
            frames = driver.find_elements(By.CSS_SELECTOR, "iframe, frame")

            for i in range(len(frames)):
                try:
                    driver.switch_to.default_content()
                    frames = driver.find_elements(By.CSS_SELECTOR, "iframe, frame")
                    driver.switch_to.frame(frames[i])

                    click_all_posts_if_exists(driver)

                    rows = driver.find_elements(By.CSS_SELECTOR, "table tbody tr")
                    if rows:
                        return True
                except Exception:
                    continue
        except Exception:
            pass

        time.sleep(1)

    driver.switch_to.default_content()
    return False


def extract_announcements_from_table(driver):
    announcements = []
    seen = set()

    rows = driver.find_elements(By.CSS_SELECTOR, "table tbody tr")

    for row in rows:
        tds = row.find_elements(By.TAG_NAME, "td")

        if len(tds) < 5:
            continue

        try:
            title_el = tds[1].find_element(By.TAG_NAME, "a")
            raw_title = " ".join(title_el.text.split())
        except Exception:
            raw_title = " ".join(tds[1].text.split())

        date_text = " ".join(tds[3].text.split())

        if not raw_title:
            continue

        course, clean_title = extract_course_and_title(raw_title)

        item = {
            "과목": course,
            "제목": clean_title,
            "등록 일시": date_text,
        }

        key = (item["과목"], item["제목"], item["등록 일시"])
        if key in seen:
            continue

        seen.add(key)
        announcements.append(item)

    return announcements


def scrape_announcements(driver):
    try:
        print("공지사항 페이지 여는 중...")
        log("공지사항 페이지 이동")
        driver.switch_to.default_content()
        driver.get(ANNOUNCEMENTS_URL)
        time.sleep(5)

        ok = switch_to_frame_with_table(driver, timeout=60)

        if not ok:
            screenshot_path = BASE_DIR / "announcement_debug.png"
            html_path = BASE_DIR / "announcement_debug.html"

            driver.save_screenshot(str(screenshot_path))
            with open(html_path, "w", encoding="utf-8") as f:
                f.write(driver.page_source)

            log("공지사항 표를 찾지 못함")
            return []

        announcements = extract_announcements_from_table(driver)
        print(f"공지사항 {len(announcements)}개를 찾았습니다.")
        log(f"공지사항 {len(announcements)}개 추출")
        return announcements

    except Exception as e:
        log(f"공지사항 수집 중 오류: {repr(e)}")
        try:
            driver.save_screenshot(str(BASE_DIR / "announcement_error.png"))
            with open(BASE_DIR / "announcement_error.html", "w", encoding="utf-8") as f:
                f.write(driver.page_source)
        except Exception:
            pass
        return []


# =========================
# 3) 캘린더 페이지 처리
# =========================
# =========================
# 3) 캘린더(.ics 피드) 처리
# =========================
CALENDAR_URL = "https://mylms.korea.ac.kr/calendar#view_name=month&view_start=2026-05-19"


def normalize_ics_url(url: str, base_url: str = "") -> str:
    url = html.unescape((url or "").strip())

    if url.startswith("webcal://"):
        url = "https://" + url[len("webcal://"):]

    if base_url:
        url = urljoin(base_url, url)

    return url


def extract_first_ics_url_from_text(text: str):
    text = html.unescape(text or "")
    text = text.replace("\\/", "/")

    patterns = [
        r"((?:https?|webcal)://[^\s\"'<>]+?\.ics[^\s\"'<>]*)",
        r"(/feeds/calendars/[^\s\"'<>]+?\.ics[^\s\"'<>]*)",
    ]

    for pattern in patterns:
        match = re.search(pattern, text)
        if match:
            return match.group(1)

    return None


def get_page_text_for_ics_search(driver):
    script = """
        const values = [];

        document.querySelectorAll("input, textarea").forEach(el => {
            values.push(el.value || "");
            values.push(el.getAttribute("value") || "");
            values.push(el.placeholder || "");
        });

        document.querySelectorAll("a").forEach(el => {
            values.push(el.href || "");
            values.push(el.textContent || "");
        });

        document.querySelectorAll("*").forEach(el => {
            const title = el.getAttribute("title");
            const aria = el.getAttribute("aria-label");
            if (title) values.push(title);
            if (aria) values.push(aria);
        });

        if (document.body) {
            values.push(document.body.innerText || "");
            values.push(document.body.innerHTML || "");
        }

        return values.join("\\n");
    """

    return driver.execute_script(script)


def click_calendar_feed_button_current_context(driver):
    """
    현재 문서/frame 안에서 '캘린더 피드' 버튼을 찾아 클릭한다.
    """
    script = """
        const candidates = Array.from(document.querySelectorAll(
            "button, a, input[type='button'], input[type='submit'], div[role='button'], span[role='button']"
        ));

        for (const el of candidates) {
            const text = [
                el.innerText,
                el.value,
                el.title,
                el.getAttribute("aria-label"),
                el.id,
                el.className
            ].filter(Boolean).join(" ").trim();

            if (
                text.includes("캘린더 피드") ||
                text.toLowerCase().includes("calendar feed")
            ) {
                el.scrollIntoView({block: "center"});
                el.click();
                return text;
            }
        }

        return null;
    """

    clicked_text = driver.execute_script(script)

    if clicked_text:
        log(f"캘린더 피드 버튼 클릭 성공: {clicked_text}")
        return True

    return False


def click_calendar_feed_button(driver, timeout=30):
    """
    기본 문서와 iframe을 돌면서 '캘린더 피드' 버튼을 클릭한다.
    """
    end_time = time.time() + timeout

    while time.time() < end_time:
        try:
            driver.switch_to.default_content()

            if click_calendar_feed_button_current_context(driver):
                return True

            frames = driver.find_elements(By.CSS_SELECTOR, "iframe, frame")

            for i in range(len(frames)):
                try:
                    driver.switch_to.default_content()
                    frames = driver.find_elements(By.CSS_SELECTOR, "iframe, frame")
                    driver.switch_to.frame(frames[i])

                    if click_calendar_feed_button_current_context(driver):
                        return True

                except Exception:
                    continue

        except Exception as e:
            log(f"캘린더 피드 버튼 탐색 중 오류: {e}")

        time.sleep(1)

    driver.switch_to.default_content()
    return False


def find_ics_url_after_click(driver, timeout=20):
    """
    캘린더 피드 버튼 클릭 후 뜨는 창에서 .ics URL을 찾는다.
    """
    end_time = time.time() + timeout

    while time.time() < end_time:
        try:
            driver.switch_to.default_content()
            page_text = get_page_text_for_ics_search(driver)
            ics_url = extract_first_ics_url_from_text(page_text)

            if ics_url:
                return normalize_ics_url(ics_url, driver.current_url)

            frames = driver.find_elements(By.CSS_SELECTOR, "iframe, frame")

            for i in range(len(frames)):
                try:
                    driver.switch_to.default_content()
                    frames = driver.find_elements(By.CSS_SELECTOR, "iframe, frame")
                    driver.switch_to.frame(frames[i])

                    page_text = get_page_text_for_ics_search(driver)
                    ics_url = extract_first_ics_url_from_text(page_text)

                    if ics_url:
                        return normalize_ics_url(ics_url, driver.current_url)

                except Exception:
                    continue

        except Exception as e:
            log(f"ICS URL 탐색 중 오류: {e}")

        time.sleep(1)

    driver.switch_to.default_content()
    return None


def download_ics_with_cookies(ics_url: str, driver) -> str:
    """
    추출한 ICS URL을 다운로드한다.
    로그인 세션이 필요한 경우를 대비해서 Selenium 쿠키도 같이 보낸다.
    """
    cookie_header = "; ".join(
        f"{cookie['name']}={cookie['value']}"
        for cookie in driver.get_cookies()
    )

    request = Request(
        ics_url,
        headers={
            "User-Agent": "Mozilla/5.0",
            "Cookie": cookie_header,
        }
    )

    with urlopen(request) as response:
        return response.read().decode("utf-8", errors="replace")


def unfold_ics_lines(text: str):
    raw_lines = text.replace("\r\n", "\n").replace("\r", "\n").split("\n")
    lines = []

    for line in raw_lines:
        if (line.startswith(" ") or line.startswith("\t")) and lines:
            lines[-1] += line[1:]
        else:
            lines.append(line)

    return lines


def unescape_ics_text(value: str) -> str:
    return (
        value.replace("\\n", "\n")
        .replace("\\N", "\n")
        .replace("\\,", ",")
        .replace("\\;", ";")
        .replace("\\\\", "\\")
    )


def parse_datetime(key_part: str, value: str):
    value = value.strip()

    if "VALUE=DATE" in key_part or (len(value) == 8 and value.isdigit()):
        try:
            return datetime.strptime(value, "%Y%m%d")
        except ValueError:
            return None

    if value.endswith("Z"):
        for fmt in ("%Y%m%dT%H%M%SZ", "%Y%m%dT%H%MZ"):
            try:
                dt = datetime.strptime(value, fmt)
                return dt.replace(tzinfo=timezone.utc)
            except ValueError:
                continue
        return None

    for fmt in ("%Y%m%dT%H%M%S", "%Y%m%dT%H%M"):
        try:
            return datetime.strptime(value, fmt)
        except ValueError:
            continue

    return None


def format_datetime(dt):
    if dt is None:
        return ""

    if dt.tzinfo is not None:
        kst = timezone(timedelta(hours=9))
        dt = dt.astimezone(kst)

    return dt.strftime("%Y-%m-%d %H:%M")


def detect_type(title: str, description: str) -> str:
    text = f"{title} {description}".lower()

    due_keywords = [
        "due", "assignment", "quiz", "exam", "test", "submission",
        "과제", "마감", "퀴즈", "시험", "제출"
    ]

    material_keywords = [
        "material", "file", "lecture note", "slides", "handout",
        "자료", "강의자료", "수업자료", "pdf", "lecture"
    ]

    if any(keyword in text for keyword in due_keywords):
        return "과제/마감"

    if any(keyword in text for keyword in material_keywords):
        return "강의자료/수업"

    return "기타 일정"


def parse_ics_events(ics_text: str):
    lines = unfold_ics_lines(ics_text)
    events = []
    current_event = None

    for line in lines:
        if line == "BEGIN:VEVENT":
            current_event = {}
            continue

        if line == "END:VEVENT":
            if current_event:
                title = unescape_ics_text(current_event.get("SUMMARY", ""))
                description = unescape_ics_text(current_event.get("DESCRIPTION", ""))
                url = current_event.get("URL", "")

                start_dt = current_event.get("_START_DT")
                end_dt = current_event.get("_END_DT")

                event_type = detect_type(title, description)

                if event_type == "과제/마감":
                    date_str = format_datetime(start_dt or end_dt)
                else:
                    date_str = format_datetime(end_dt or start_dt)

                events.append({
                    "유형": event_type,
                    "제목": title,
                    "종료/마감": date_str,
                    "설명": description,
                    "URL": url
                })

            current_event = None
            continue

        if current_event is None:
            continue

        if ":" not in line:
            continue

        key_part, value = line.split(":", 1)
        key_name = key_part.split(";", 1)[0].upper()

        if key_name == "DTSTART":
            current_event["_START_DT"] = parse_datetime(key_part, value)
        elif key_name == "DTEND":
            current_event["_END_DT"] = parse_datetime(key_part, value)
        else:
            current_event[key_name] = value

    def sort_key(event):
        date_str = event["종료/마감"]

        if not date_str:
            return datetime.max

        for fmt in ("%Y-%m-%d %H:%M", "%Y-%m-%d"):
            try:
                return datetime.strptime(date_str, fmt)
            except ValueError:
                continue

        return datetime.max

    events.sort(key=sort_key)
    return events


def scrape_calendar(driver):
    try:
        print("캘린더 페이지 여는 중...")
        log("캘린더 페이지 이동")

        driver.switch_to.default_content()
        driver.get(CALENDAR_URL)
        time.sleep(6)

        clicked = click_calendar_feed_button(driver, timeout=30)

        if not clicked:
            driver.save_screenshot(str(BASE_DIR / "calendar_feed_button_debug.png"))
            with open(BASE_DIR / "calendar_feed_button_debug.html", "w", encoding="utf-8") as f:
                f.write(driver.page_source)

            log("캘린더 피드 버튼을 찾지 못함")
            return []

        time.sleep(2)

        ics_url = find_ics_url_after_click(driver, timeout=20)

        if not ics_url:
            driver.save_screenshot(str(BASE_DIR / "calendar_ics_url_debug.png"))
            with open(BASE_DIR / "calendar_ics_url_debug.html", "w", encoding="utf-8") as f:
                f.write(driver.page_source)

            log("ICS URL을 찾지 못함")
            return []

        log(f"ICS URL 추출 성공: {ics_url}")

        print("ICS 피드 다운로드 중...")
        ics_text = download_ics_with_cookies(ics_url, driver)

        print("ICS 일정 파싱 중...")
        events = parse_ics_events(ics_text)

        print(f"캘린더 일정 {len(events)}개를 찾았습니다.")
        log(f"캘린더 일정 {len(events)}개 추출")

        return events

    except Exception as e:
        log(f"캘린더 수집 중 오류: {repr(e)}")

        try:
            driver.save_screenshot(str(BASE_DIR / "calendar_error.png"))
            with open(BASE_DIR / "calendar_error.html", "w", encoding="utf-8") as f:
                f.write(driver.page_source)
        except Exception:
            pass

        return []


# =========================
# 4) 저장 처리
# =========================
def save_calendar_to_csv(events, filename):
    if not events:
        return

    fieldnames = ["유형", "제목", "종료/마감", "설명", "URL"]

    with open(filename, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(events)


def style_sheet(ws, headers, widths):
    from openpyxl.styles import Font, PatternFill, Alignment

    ws.append(headers)

    header_fill = PatternFill("solid", fgColor="D9EAF7")
    header_font = Font(bold=True)

    for col_idx, _ in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

# =========================
# 우선순위 정렬 처리
# =========================
def normalize_text(value):
    return re.sub(r"\s+", " ", str(value or "")).strip().lower()


def split_keywords(value):
    keywords = re.split(r"[,;/\n]+", str(value or ""))
    return [normalize_text(k) for k in keywords if normalize_text(k)]


def to_weight(value, default=1):
    try:
        n = int(value)
    except Exception:
        n = default

    if n < 1:
        n = 1
    if n > 5:
        n = 5

    return n


def is_enabled(value):
    return str(value or "").strip().upper() == "Y"


def get_type_settings(settings):
    type_settings = settings.get("type_settings", []) if settings else []

    if type_settings:
        return type_settings

    return [
        {"type": "과제", "enabled": "Y", "weight": 5, "keywords": "과제, 제출, 마감, assignment, due, homework"},
        {"type": "시험", "enabled": "Y", "weight": 5, "keywords": "시험, exam, test"},
        {"type": "중간", "enabled": "Y", "weight": 5, "keywords": "중간, midterm"},
        {"type": "기말", "enabled": "Y", "weight": 5, "keywords": "기말, final"},
        {"type": "퀴즈", "enabled": "Y", "weight": 4, "keywords": "퀴즈, quiz"},
        {"type": "발표", "enabled": "Y", "weight": 4, "keywords": "발표, presentation"},
        {"type": "강의자료", "enabled": "N", "weight": 2, "keywords": "강의자료, 자료, pdf, lecture, slide, handout"},
        {"type": "공지", "enabled": "Y", "weight": 1, "keywords": "공지, 안내, notice"},
        {"type": "기타", "enabled": "N", "weight": 1, "keywords": "기타"},
    ]


def get_course_settings(settings):
    if not settings:
        return []

    return settings.get("course_settings", [])


def find_type_match(text, default_type, settings):
    text_norm = normalize_text(text)
    default_norm = normalize_text(default_type)

    matches = []

    for cfg in get_type_settings(settings):
        type_name = str(cfg.get("type", "")).strip()
        keywords = [normalize_text(type_name)] + split_keywords(cfg.get("keywords", ""))

        matched_keywords = []

        for keyword in keywords:
            if keyword and keyword in text_norm:
                matched_keywords.append(keyword)

        if matched_keywords:
            matches.append({
                "config": cfg,
                "matched_keywords": matched_keywords,
                "weight": to_weight(cfg.get("weight", 1)),
                "longest_keyword": max(len(k) for k in matched_keywords),
            })

    if matches:
        matches.sort(
            key=lambda x: (x["weight"], x["longest_keyword"]),
            reverse=True
        )
        best = matches[0]
        return best["config"], best["matched_keywords"]

    # 텍스트에서 직접 못 찾으면 기존 캘린더 유형 또는 공지사항 기본 유형으로 판단
    for cfg in get_type_settings(settings):
        type_name = normalize_text(cfg.get("type", ""))

        if type_name and type_name in default_norm:
            return cfg, [type_name]

    # 공지사항은 기본적으로 공지로 처리
    if default_norm == "공지":
        for cfg in get_type_settings(settings):
            if normalize_text(cfg.get("type", "")) == "공지":
                return cfg, ["공지"]

    # 그래도 없으면 기타
    for cfg in get_type_settings(settings):
        if normalize_text(cfg.get("type", "")) == "기타":
            return cfg, ["기타"]

    return None, []


def find_course_match(text, settings):
    text_norm = normalize_text(text)

    matched_enabled = []
    matched_disabled = []

    for cfg in get_course_settings(settings):
        course = str(cfg.get("course", "")).strip()
        course_norm = normalize_text(course)

        if not course_norm:
            continue

        if course_norm in text_norm:
            item = {
                "course": course,
                "weight": to_weight(cfg.get("weight", 1)),
                "enabled": is_enabled(cfg.get("enabled", "Y")),
            }

            if item["enabled"]:
                matched_enabled.append(item)
            else:
                matched_disabled.append(item)

    # Y로 설정된 과목/키워드가 하나라도 매칭되면 그중 가중치 가장 높은 것 사용
    if matched_enabled:
        matched_enabled.sort(key=lambda x: x["weight"], reverse=True)
        best = matched_enabled[0]
        return best["course"], best["weight"], False

    # Y 매칭은 없고 N 매칭만 있으면 제외
    if matched_disabled:
        matched_disabled.sort(key=lambda x: x["weight"], reverse=True)
        best = matched_disabled[0]
        return best["course"], best["weight"], True

    # 아무 과목 키워드도 안 맞으면 제외하지 않고 과목 가중치 0
    return "", 0, False


def build_priority_rows(calendar_events, announcements, settings):
    rows = []

    # 1) 캘린더 일정 우선순위화
    for event in calendar_events:
        title = event.get("제목", "")
        description = event.get("설명", "")
        default_type = event.get("유형", "기타")
        date_text = event.get("종료/마감", "")
        url = event.get("URL", "")

        full_text = f"{default_type} {title} {description}"

        type_cfg, matched_type_keywords = find_type_match(
            full_text,
            default_type,
            settings
        )

        if not type_cfg:
            continue

        if not is_enabled(type_cfg.get("enabled", "N")):
            continue

        course_name, course_weight, exclude_by_course = find_course_match(
            full_text,
            settings
        )

        if exclude_by_course:
            continue

        type_name = type_cfg.get("type", default_type)
        type_weight = to_weight(type_cfg.get("weight", 1))
        score = type_weight + course_weight

        rows.append({
            "점수": score,
            "출처": "캘린더",
            "유형": type_name,
            "과목/키워드": course_name,
            "제목": title,
            "날짜": date_text,
            "매칭 키워드": ", ".join(matched_type_keywords),
            "URL": url,
        })

    # 2) 공지사항 우선순위화
    for item in announcements:
        course = item.get("과목", "")
        title = item.get("제목", "")
        date_text = item.get("등록 일시", "")

        full_text = f"공지 {course} {title}"

        type_cfg, matched_type_keywords = find_type_match(
            full_text,
            "공지",
            settings
        )

        if not type_cfg:
            continue

        if not is_enabled(type_cfg.get("enabled", "N")):
            continue

        course_name, course_weight, exclude_by_course = find_course_match(
            full_text,
            settings
        )

        if exclude_by_course:
            continue

        if not course_name:
            course_name = course

        type_name = type_cfg.get("type", "공지")
        type_weight = to_weight(type_cfg.get("weight", 1))
        score = type_weight + course_weight

        rows.append({
            "점수": score,
            "출처": "공지사항",
            "유형": type_name,
            "과목/키워드": course_name,
            "제목": title,
            "날짜": date_text,
            "매칭 키워드": ", ".join(matched_type_keywords),
            "URL": "",
        })

    rows.sort(key=lambda x: (-x["점수"], str(x["날짜"])))
    return rows

def save_to_xlsx(calendar_events, announcements, filename, settings=None):
    try:
        from openpyxl import Workbook
    except ImportError:
        print("openpyxl이 설치되지 않아 XLSX 저장은 건너뜁니다.")
        print("설치 명령어: pip install openpyxl")
        return

    wb = Workbook()

    # =========================
    # 1. LMS 일정 시트
    # =========================
    ws1 = wb.active
    ws1.title = "LMS 일정"

    event_headers = ["유형", "제목", "종료/마감", "설명", "URL"]
    style_sheet(
        ws1,
        event_headers,
        {"A": 14, "B": 45, "C": 18, "D": 80, "E": 60},
    )

    for event in calendar_events:
        ws1.append([
            event.get("유형", ""),
            event.get("제목", ""),
            event.get("종료/마감", ""),
            event.get("설명", ""),
            event.get("URL", ""),
        ])

    # =========================
    # 2. 공지사항 시트
    # =========================
    ws2 = wb.create_sheet("공지사항")

    ann_headers = ["과목", "제목", "등록 일시"]
    style_sheet(
        ws2,
        ann_headers,
        {"A": 45, "B": 90, "C": 22},
    )

    for item in announcements:
        ws2.append([
            item.get("과목", ""),
            item.get("제목", ""),
            item.get("등록 일시", ""),
        ])

    # =========================
    # 3. 우선순위 정렬 시트
    # =========================
    ws3 = wb.create_sheet("우선순위 정렬")

    priority_headers = [
        "점수",
        "출처",
        "유형",
        "과목/키워드",
        "제목",
        "날짜",
        "매칭 키워드",
        "URL",
    ]

    style_sheet(
        ws3,
        priority_headers,
        {
            "A": 10,
            "B": 12,
            "C": 14,
            "D": 28,
            "E": 90,
            "F": 22,
            "G": 35,
            "H": 60,
        },
    )

    priority_rows = build_priority_rows(calendar_events, announcements, settings)

    for row in priority_rows:
        ws3.append([
            row["점수"],
            row["출처"],
            row["유형"],
            row["과목/키워드"],
            row["제목"],
            row["날짜"],
            row["매칭 키워드"],
            row["URL"],
        ])

    wb.save(filename)
    print(f"XLSX 저장 완료: {filename}")
    print(f"우선순위 정렬 항목: {len(priority_rows)}개")


# =========================
# 5) 실행
# =========================
def main():
    DEBUG_FILE.write_text("", encoding="utf-8")

    data = read_payload()
    portal_url = data["portal_url"]
    user_id = data["user_id"]
    password = data["password"]
    settings = data.get("settings", {})

    chrome_options = Options()
    chrome_options.add_argument("--start-maximized")
    chrome_options.add_experimental_option("detach", KEEP_BROWSER_OPEN)

    driver = webdriver.Chrome(options=chrome_options)

    try:
        print("브라우저 시작")
        log("브라우저 시작")
        log(f"접속 URL: {portal_url}")

        # 1. 로그인 페이지 접속
        driver.get(portal_url)
        time.sleep(3)

        # 2. 아이디/비밀번호 입력 + 로그인 버튼 클릭
        driver.switch_to.default_content()

        login_success = recursive_find_and_fill(driver, user_id, password)

        if not login_success:
            raise RuntimeError(
                "로그인 입력칸을 찾지 못했습니다. login_debug.txt 파일을 확인해주세요."
            )

        # 3. 로그인 후 LMS 이동 대기
        wait_after_login(driver, timeout=50)

        # 4. 공지사항 수집
        announcements = scrape_announcements(driver)

        # 5. 캘린더 페이지 접속 → 캘린더 피드 버튼 클릭 → ICS 링크 추출 → 일정 파싱
        calendar_events = scrape_calendar(driver)

        # 6. 저장
        save_calendar_to_csv(calendar_events, OUTPUT_CSV)
        save_to_xlsx(calendar_events, announcements, OUTPUT_XLSX, settings)

        print("완료!")
        print(f"캘린더 일정: {len(calendar_events)}개")
        print(f"공지사항: {len(announcements)}개")
        print(f"저장 위치: {OUTPUT_XLSX}")

    except Exception as e:
        log("")
        log("ERROR")
        log(str(e))

        print("로그인/수집 자동화 중 오류가 발생했습니다.", file=sys.stderr)
        print(str(e), file=sys.stderr)
        print(f"디버그 파일 위치: {DEBUG_FILE}", file=sys.stderr)
        sys.exit(1)

    finally:
        if not KEEP_BROWSER_OPEN:
            try:
                driver.quit()
            except Exception:
                pass


if __name__ == "__main__":
    main()