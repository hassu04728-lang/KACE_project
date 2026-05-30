import imaplib
import email
from email.header import decode_header
from email.utils import parsedate_to_datetime
import hashlib
import json
import os
import re
import sys
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

IMAP_SERVER = "imap.worksmobile.com"
EMAIL_USER = "hassu04@korea.ac.kr"

HEADERS = [
    "Module", "Source", "Category", "Title", "Summary", "PostedDate",
    "Deadline", "Status", "Key", "Weight", "LastSeen"
]


def load_env_from_parent():
    """KACE/.env에 있는 NAVERWORKS_APP_PASSWORD를 환경변수처럼 읽기 위한 보조 함수."""
    candidates = [
        Path(__file__).resolve().parent.parent / ".env",
        Path.cwd() / ".env",
    ]
    for env_path in candidates:
        if not env_path.exists():
            continue
        try:
            for line in env_path.read_text(encoding="utf-8").splitlines():
                line = line.strip()
                if not line or line.startswith("#") or "=" not in line:
                    continue
                key, value = line.split("=", 1)
                key = key.strip()
                value = value.strip().strip('"').strip("'")
                os.environ.setdefault(key, value)
        except Exception:
            pass


def read_payload():
    raw = sys.stdin.read().strip()
    if not raw:
        base_dir = Path(__file__).resolve().parent.parent
        return {
            "output_path": str(base_dir / "temp" / "mail_result.xlsx"),
            "max_mails": 50,
            "keywords": [
                {"keyword": "장학", "weight": 5},
                {"keyword": "인턴", "weight": 3},
                {"keyword": "공지", "weight": 1},
            ],
        }
    return json.loads(raw)


def clean_text(text):
    if not text:
        return ""
    text = re.sub(r"<[^>]+>", " ", text)
    text = text.replace("&nbsp;", " ")
    text = re.sub(r"\s+", " ", text)
    return text.strip()


def decode_mime_header(value):
    if not value:
        return ""
    decoded_parts = decode_header(value)
    result = ""
    for part, encoding in decoded_parts:
        if isinstance(part, bytes):
            result += part.decode(encoding or "utf-8", errors="ignore")
        else:
            result += part
    return result


def extract_date_from_title(title):
    pattern = r"(\d{4}[-./]\d{1,2}[-./]\d{1,2}|\d{1,2}[./월]\s?\d{1,2}(?:일)?)"
    match = re.search(pattern, title or "")
    if match:
        return match.group(1)
    return "본문 확인 필요"


def extract_body(msg):
    body = ""
    if msg.is_multipart():
        plain_body = ""
        html_body = ""
        for part in msg.walk():
            content_type = part.get_content_type()
            charset = part.get_content_charset() or "utf-8"
            try:
                payload = part.get_payload(decode=True)
                if not payload:
                    continue
                decoded = payload.decode(charset, errors="ignore")
                if content_type == "text/plain" and not plain_body:
                    plain_body = decoded
                elif content_type == "text/html" and not html_body:
                    html_body = decoded
            except Exception:
                continue
        body = plain_body if plain_body else html_body
    else:
        charset = msg.get_content_charset() or "utf-8"
        payload = msg.get_payload(decode=True)
        if payload:
            body = payload.decode(charset, errors="ignore")
    return clean_text(body)


def build_keyword_map(payload):
    items = payload.get("keywords", [])
    keyword_map = {}
    for item in items:
        keyword = str(item.get("keyword", "")).strip()
        if not keyword:
            continue
        try:
            weight = int(item.get("weight", 1))
        except Exception:
            weight = 1
        weight = max(1, min(5, weight))
        keyword_map[keyword] = weight
    if not keyword_map:
        keyword_map = {"장학": 5, "인턴": 3, "공지": 1}
    return keyword_map


def collect_mails(keyword_map, max_mails=50):
    email_pass = os.getenv("NAVERWORKS_APP_PASSWORD")
    if not email_pass:
        raise ValueError(".env 파일에 NAVERWORKS_APP_PASSWORD를 설정해야 합니다.")

    keywords = list(keyword_map.keys())
    print(f"[*] 활성화된 메일 키워드 및 가중치: {keyword_map}")
    print("[*] 네이버웍스 IMAP 접속 중...")

    mail = imaplib.IMAP4_SSL(IMAP_SERVER, 993)
    mail.login(EMAIL_USER, email_pass)
    mail.select("INBOX")

    results = []
    status, messages = mail.search(None, "ALL")
    target_ids = messages[0].split()[-int(max_mails):]

    print(f"[*] 최근 {len(target_ids)}개 메일 검사 중...")

    for m_id in target_ids:
        _, msg_data = mail.fetch(m_id, "(RFC822)")
        for response_part in msg_data:
            if not isinstance(response_part, tuple):
                continue
            msg = email.message_from_bytes(response_part[1])
            subject = decode_mime_header(msg.get("Subject"))
            if not subject:
                continue

            matched = [kw for kw in keywords if str(kw) in subject]
            if not matched:
                continue

            kw = matched[0]
            weight = keyword_map.get(kw, 1)
            body = extract_body(msg)
            deadline_info = extract_date_from_title(subject)

            mail_date_raw = msg.get("Date")
            try:
                posted_date = parsedate_to_datetime(mail_date_raw).strftime("%Y-%m-%d %H:%M")
            except Exception:
                posted_date = mail_date_raw or ""

            unique_key = hashlib.md5(f"{subject}{mail_date_raw}".encode("utf-8")).hexdigest()
            module = "C" if any(ck in str(kw) for ck in ["채용", "인턴", "반도체", "공정"]) else "B"

            results.append({
                "Module": module,
                "Source": "Email",
                "Category": kw,
                "Title": subject.strip(),
                "Summary": body[:150],
                "PostedDate": posted_date,
                "Deadline": deadline_info,
                "Status": "미확인",
                "Key": unique_key,
                "Weight": weight,
                "LastSeen": datetime.now().strftime("%Y-%m-%d"),
            })
            print(f" -> 수집 성공: {subject[:25]}... / 키워드: {kw} / 가중치: {weight}")

    mail.logout()

    # 같은 실행 중 중복 제거
    seen = set()
    unique_results = []
    for item in results:
        key = item.get("Key")
        if key in seen:
            continue
        seen.add(key)
        unique_results.append(item)

    print(f"[*] 검사한 메일: {len(target_ids)}개 / 키워드 매칭: {len(unique_results)}개")
    return unique_results


def write_result_xlsx(output_path, rows):
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "DB_Raw"

    header_fill = PatternFill("solid", fgColor="D9EAF7")
    for col_idx, header in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row_idx, row in enumerate(rows, start=2):
        for col_idx, header in enumerate(HEADERS, start=1):
            ws.cell(row=row_idx, column=col_idx, value=row.get(header, ""))

    ws.freeze_panes = "A2"
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    widths = {
        "A": 10, "B": 12, "C": 14, "D": 55, "E": 60,
        "F": 18, "G": 18, "H": 12, "I": 34, "J": 10, "K": 14,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    wb.save(output_path)
    print(f"[*] 결과 파일 저장 완료: {output_path}")


def main():
    print("메일 수집 코드 실행 시작")
    load_env_from_parent()
    payload = read_payload()
    output_path = payload.get("output_path") or str(Path(__file__).resolve().parent.parent / "temp" / "mail_result.xlsx")
    max_mails = int(payload.get("max_mails", 50))
    keyword_map = build_keyword_map(payload)
    rows = collect_mails(keyword_map, max_mails=max_mails)
    write_result_xlsx(output_path, rows)
    print("메일 수집 완료")


if __name__ == "__main__":
    main()
