import os
import sys
import json
import html
import re
from pathlib import Path
from collections import Counter
from email.utils import parsedate_to_datetime

try:
    import requests
except ImportError:
    raise ImportError("requests가 설치되어 있지 않습니다. 터미널에서 python -m pip install requests 를 실행하세요.")

try:
    from dotenv import load_dotenv
except ImportError:
    raise ImportError("python-dotenv가 설치되어 있지 않습니다. 터미널에서 python -m pip install python-dotenv 를 실행하세요.")

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
except ImportError:
    raise ImportError("openpyxl이 설치되어 있지 않습니다. 터미널에서 python -m pip install openpyxl 를 실행하세요.")


NAVER_NEWS_URL = "https://openapi.naver.com/v1/search/news.json"

SEARCH_QUERY_MAP = {
    "반도체": "반도체",
    "철강": "철강 OR 금속",
    "채용": "채용 OR 공채 OR 인턴",
    "세라믹": "세라믹",
    "고분자": "고분자",
    "디스플레이": "디스플레이",
    "고려대학교": "고려대학교",
}

STOPWORDS = {
    "있다", "없다", "대한", "위한", "통해", "관련", "기자", "뉴스", "기사",
    "이번", "최근", "이날", "지난", "오늘", "오전", "오후", "시장", "산업",
    "분야", "기업", "국내", "글로벌", "정부", "중심", "확대", "추진", "발표",
    "지원", "전망", "증가", "감소", "변화", "진행", "기반", "정도", "경우",
    "채용", "공채", "인턴"
}


def read_payload() -> dict:
    raw = sys.stdin.read().strip()
    if not raw:
        raise RuntimeError("VBA에서 전달된 설정값이 없습니다.")
    return json.loads(raw)


def load_env_files():
    script_dir = Path(__file__).resolve().parent
    kace_dir = script_dir.parent

    # KACE\.env 우선, 그 다음 현재 작업 폴더 .env
    for env_path in [kace_dir / ".env", Path.cwd() / ".env"]:
        if env_path.exists():
            load_dotenv(env_path)
            print(f"[*] .env 로드: {env_path}")
            return

    # 파일이 없어도 환경변수에 이미 들어있을 수 있으므로 load_dotenv만 호출
    load_dotenv()
    print("[!] .env 파일을 찾지 못했습니다. 환경변수에서 API 키를 확인합니다.")


def clean_html_text(text: str) -> str:
    if not text:
        return ""

    text = html.unescape(text)
    text = text.replace("<b>", "").replace("</b>", "")
    text = re.sub(r"<[^>]+>", " ", text)
    text = re.sub(r"\s+", " ", text)
    return text.strip()


def parse_pubdate(pub_date: str) -> str:
    try:
        dt = parsedate_to_datetime(pub_date)
        return dt.strftime("%Y-%m-%d %H:%M")
    except Exception:
        return pub_date or ""


def normalize_keywords(keyword_items: list[dict]) -> list[dict]:
    active = []

    for item in keyword_items:
        keyword = str(item.get("keyword") or item.get("키워드") or "").strip()
        search_query = str(
            item.get("search_query")
            or item.get("검색어")
            or SEARCH_QUERY_MAP.get(keyword, keyword)
        ).strip()

        if keyword:
            active.append({
                "키워드": keyword,
                "검색어": search_query if search_query else keyword,
            })

    if not active:
        raise ValueError("사용여부가 Y인 뉴스 키워드가 없습니다. Settings 시트의 뉴스 키워드 영역을 확인하세요.")

    return active


def fetch_news_by_query(keyword: str, search_query: str, client_id: str, client_secret: str,
                        display: int = 10, start: int = 1, sort: str = "date",
                        timeout: int = 10) -> list[dict]:
    headers = {
        "X-Naver-Client-Id": client_id,
        "X-Naver-Client-Secret": client_secret,
    }

    params = {
        "query": search_query,
        "display": display,
        "start": start,
        "sort": sort,
    }

    response = requests.get(
        NAVER_NEWS_URL,
        headers=headers,
        params=params,
        timeout=timeout,
    )
    response.raise_for_status()

    data = response.json()
    items = data.get("items", [])

    rows = []
    for idx, item in enumerate(items, start=1):
        title = clean_html_text(item.get("title", ""))
        description = clean_html_text(item.get("description", ""))
        pub_date = parse_pubdate(item.get("pubDate", ""))
        originallink = item.get("originallink", "")
        naver_link = item.get("link", "")
        final_link = originallink if originallink else naver_link

        rows.append({
            "키워드": keyword,
            "검색어": search_query,
            "순위": idx,
            "기사제목": title,
            "날짜": pub_date,
            "요약": description,
            "기사링크": final_link,
        })

    return rows


def tokenize(text: str) -> list[str]:
    if not text:
        return []

    text = re.sub(r"[^0-9A-Za-z가-힣\s]", " ", text)
    words = re.findall(r"[0-9A-Za-z가-힣]{2,}", text)

    result = []
    for word in words:
        if word in STOPWORDS:
            continue
        if word.isdigit():
            continue
        result.append(word)

    return result


def build_trend_paragraph(keyword: str, search_query: str, news_rows: list[dict]) -> str:
    if not news_rows:
        return f"{keyword} 관련 기사가 수집되지 않아 최신 동향을 정리하지 못했다."

    combined = " ".join(
        f"{row.get('기사제목', '')} {row.get('요약', '')}"
        for row in news_rows
    )

    tokens = tokenize(combined)
    counter = Counter(tokens)
    top_terms = [word for word, _ in counter.most_common(6)]

    if not top_terms:
        return (
            f"{keyword} 관련 상위 기사들을 종합하면, 최근에는 해당 분야에서 정책 변화, "
            f"시장 흐름, 기술 개발 및 기관·기업의 움직임이 함께 나타나고 있다. "
            f"단기 이슈보다는 향후 방향성과 실제 적용 가능성을 보여주는 기사들이 주로 노출되는 경향이 있다."
        )

    term_text = ", ".join(top_terms[:5])

    return (
        f"{keyword} 관련 상위 기사들을 종합하면 최근 흐름은 {term_text} 등을 중심으로 형성되고 있다. "
        f"전반적으로는 단순 사건성 보도보다 기술 개발, 투자·사업 확대, 제도 변화, 기관 및 기업의 전략 움직임이 "
        f"함께 나타나는 모습이며, 이를 통해 {keyword} 분야가 현재 어떤 주제에 관심이 집중되고 있는지 확인할 수 있다. "
        f"이번 요약은 검색어 '{search_query}'로 수집한 기사 제목과 요약문을 바탕으로 정리한 결과다."
    )


def style_header(ws):
    fill = PatternFill("solid", fgColor="D9EAF7")
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center", vertical="center")


def format_sheet(ws, widths: dict[str, int]):
    ws.freeze_panes = "A2"

    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def write_output_workbook(output_path: str, news_rows: list[dict], trend_rows: list[dict]):
    output = Path(output_path)
    output.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()

    ws_news = wb.active
    ws_news.title = "뉴스 수집"
    news_headers = ["키워드", "검색어", "순위", "기사제목", "날짜", "요약", "기사링크"]
    ws_news.append(news_headers)
    style_header(ws_news)

    for row in news_rows:
        ws_news.append([
            row.get("키워드", ""),
            row.get("검색어", ""),
            row.get("순위", ""),
            row.get("기사제목", ""),
            row.get("날짜", ""),
            row.get("요약", ""),
            row.get("기사링크", ""),
        ])

    format_sheet(ws_news, {
        "A": 16,
        "B": 24,
        "C": 8,
        "D": 55,
        "E": 18,
        "F": 70,
        "G": 70,
    })

    ws_trend = wb.create_sheet("트렌드 요약")
    trend_headers = ["키워드", "검색어", "기사건수", "종합 트렌드/최신 동향"]
    ws_trend.append(trend_headers)
    style_header(ws_trend)

    for row in trend_rows:
        ws_trend.append([
            row.get("키워드", ""),
            row.get("검색어", ""),
            row.get("기사건수", ""),
            row.get("종합 트렌드/최신 동향", ""),
        ])

    format_sheet(ws_trend, {
        "A": 16,
        "B": 24,
        "C": 10,
        "D": 90,
    })

    wb.save(output)


def main():
    payload = read_payload()
    output_path = payload.get("output_path")
    keyword_items = payload.get("keywords", [])
    display = int(payload.get("display", 10))

    if not output_path:
        raise ValueError("output_path가 전달되지 않았습니다.")

    load_env_files()

    client_id = os.getenv("NAVER_CLIENT_ID")
    client_secret = os.getenv("NAVER_CLIENT_SECRET")

    if not client_id or not client_secret:
        raise ValueError(".env 파일에 NAVER_CLIENT_ID와 NAVER_CLIENT_SECRET을 입력해야 합니다.")

    active_items = normalize_keywords(keyword_items)

    all_news_rows = []
    all_trend_rows = []

    print("[*] 선택된 뉴스 키워드")
    for item in active_items:
        print(f" - {item['키워드']} → 검색어: {item['검색어']}")

    for item in active_items:
        keyword = item["키워드"]
        search_query = item["검색어"]

        print(f"\n[수집 시작] {keyword}")

        try:
            news_rows = fetch_news_by_query(
                keyword=keyword,
                search_query=search_query,
                client_id=client_id,
                client_secret=client_secret,
                display=display,
                start=1,
                sort="date",
            )

            all_news_rows.extend(news_rows)

            trend_text = build_trend_paragraph(keyword, search_query, news_rows)
            all_trend_rows.append({
                "키워드": keyword,
                "검색어": search_query,
                "기사건수": len(news_rows),
                "종합 트렌드/최신 동향": trend_text,
            })

            print(f"[완료] {keyword}: {len(news_rows)}건 수집")

        except Exception as e:
            print(f"[오류] {keyword}: {e}")
            all_trend_rows.append({
                "키워드": keyword,
                "검색어": search_query,
                "기사건수": 0,
                "종합 트렌드/최신 동향": f"오류로 인해 수집 실패: {e}",
            })

    write_output_workbook(output_path, all_news_rows, all_trend_rows)

    print(f"\n[완료] 뉴스 수집 결과 저장: {output_path}")
    print(f"뉴스 기사 수: {len(all_news_rows)}건")
    print(f"트렌드 요약 수: {len(all_trend_rows)}건")


if __name__ == "__main__":
    main()
